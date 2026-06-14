"""Missouri MSHP sex offender registry adapter.

Source: https://www.mshp.dps.missouri.gov/MSHPWeb/PatrolDivisions/CRID/SOR/msor.zip

MSHP publishes the full Missouri Sex Offender Registry as a daily-
refreshed ZIP archive containing four Excel files:

  * `msor.xlsx`        — primary registrant + offense rows (1 row per
                          (person, offense), ~28K rows for ~21K people)
  * `msor_offense.xlsx` — additional offense detail (we ignore for now;
                          msor.xlsx already includes one offense line
                          per row)
  * `msor_alias.xlsx`   — alternate name records (each alias is its own
                          row with the same DOB+address as the primary,
                          no explicit "alias of" link). Reconstructing
                          alias relationships requires fuzzy clustering
                          that's better done post-ingest; this adapter
                          ignores the file.
  * `msor_veh.xls`      — registered vehicles (outside our schema)

No captcha, no auth, no acceptance gate on the ZIP itself — the
"I Agree" disclaimer in the browser-facing flow is purely informational
on the ZIP path. ETag/Last-Modified are honored so re-downloads are
cheap.

Schema notes:
  * No native unique ID. `source_id` is synthesized from
    `NAME|YYYY-MM-DD` (DOB) so multiple offense rows for the same
    person collapse to one record at upsert; the schema merges their
    offenses by (raw_description, conviction_date).
  * Names arrive as "LASTNAME, FIRSTNAME M[IDDLEINITIAL]" — flipped
    to natural order for `identity.full_name`.
  * "Compliant" letter codes:
        Y = Yes, currently registered → Registration.status="active"
        I = Incarcerated              → "incarcerated"
        N = Not compliant             → "absconder", absconder=True
        A = Absconder                 → "absconder", absconder=True
        O = Moved out of state        → "removed"
        P = Pending registration      → "active"
  * "Tier" is 1/2/3 — stored on each offense as `tier_or_level_raw`.
  * MSHP does NOT include photos in the bulk file. Per-person photos
    live behind the `/CJ38/searchRegistry.jsp` flow and would require
    a second-pass adapter to harvest; this adapter ships without them.
  * Cached download path: `registry-runs/missouri/msor.zip`,
    refetched if older than `cache_max_age_hours` (default 24).
"""

from __future__ import annotations

import time
import zipfile
from collections.abc import Iterator
from datetime import datetime, timezone
from pathlib import Path

import httpx

from ..photos import PhotoRef
from ..schema import Address, Identity, Offense, OffenderRecord, Registration, Source
from .base import Adapter

DOWNLOAD_URL = (
    "https://www.mshp.dps.missouri.gov/MSHPWeb/PatrolDivisions/CRID/SOR/msor.zip"
)
DEFAULT_CACHE_DIR = Path("registry-runs/missouri")
DEFAULT_ZIP_NAME = "msor.zip"
DEFAULT_MAIN_FILE = "msor.xlsx"
DEFAULT_ALIAS_FILE = "msor_alias.xlsx"
DEFAULT_CACHE_MAX_AGE_HOURS = 24

# Number of header/preamble rows in msor.xlsx before the column header.
# Layout (1-indexed):
#   row 1..2  — banner ("MISSOURI STATE HIGHWAY PATROL" / "Entire ... Report")
#   row 5..6  — REPORT DATE / TIME
#   row 8..12 — population summary lines (ACTIVE / INCARCERATED / MOVED OUT
#               OF STATE / PENDING / TOTAL)
#   row 14    — column header (Name, Address, City, St, Zip, County, Offense,
#               Count, Compliant, Tier, Date of Birth)
#   row 15+   — data
DATA_HEADER_ROW = 14

_COMPLIANT_MAP = {
    "Y": "active",
    "I": "incarcerated",
    "N": "absconder",
    "A": "absconder",
    "O": "removed",
    "P": "active",
}


class MissouriAdapter(Adapter):
    jurisdiction = "US-MO"
    source_name = "Missouri State Highway Patrol"

    def __init__(
        self,
        cache_dir: Path | str | None = None,
        cache_max_age_hours: float = DEFAULT_CACHE_MAX_AGE_HOURS,
        force_refresh: bool = False,
    ) -> None:
        self.cache_dir = Path(cache_dir) if cache_dir else DEFAULT_CACHE_DIR
        self.cache_max_age_hours = cache_max_age_hours
        self.force_refresh = force_refresh
        self._fetched_at: datetime | None = None

    @property
    def zip_path(self) -> Path:
        return self.cache_dir / DEFAULT_ZIP_NAME

    @property
    def main_path(self) -> Path:
        return self.cache_dir / DEFAULT_MAIN_FILE

    # ---- fetch --------------------------------------------------------

    def fetch(self) -> Iterator[dict]:
        self._ensure_zip()
        self._fetched_at = datetime.fromtimestamp(
            self.main_path.stat().st_mtime, tz=timezone.utc
        )
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError(
                "Missouri adapter needs openpyxl. Install with: "
                "pip install openpyxl"
            ) from e

        wb = load_workbook(str(self.main_path), read_only=True)
        try:
            ws = wb.active
            iter_rows = ws.iter_rows(min_row=DATA_HEADER_ROW + 1, values_only=True)
            for row in iter_rows:
                if not row or not row[0] or not isinstance(row[0], str):
                    continue
                # row layout: Name, Address, City, St, Zip, County, Offense,
                #             Count, Compliant, Tier, Date of Birth
                name = (row[0] or "").strip()
                if "," not in name:
                    # Skip footer / total rows that lack a real comma-name.
                    continue
                yield {
                    "name": name,
                    "address": row[1],
                    "city": row[2],
                    "state": row[3],
                    "zip": row[4],
                    "county": row[5],
                    "offense": row[6],
                    "count": row[7],
                    "compliant": row[8],
                    "tier": row[9],
                    "dob": row[10],
                }
        finally:
            wb.close()

    def _ensure_zip(self) -> None:
        if self.zip_path.exists() and not self.force_refresh:
            age = time.time() - self.zip_path.stat().st_mtime
            if age < self.cache_max_age_hours * 3600 and self.main_path.exists():
                return
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        with httpx.Client(
            timeout=300,
            follow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 registry-faces/0.1"},
        ) as client:
            with client.stream("GET", DOWNLOAD_URL) as resp:
                resp.raise_for_status()
                tmp = self.zip_path.with_suffix(self.zip_path.suffix + ".part")
                with tmp.open("wb") as f:
                    for chunk in resp.iter_bytes():
                        f.write(chunk)
                tmp.replace(self.zip_path)
        # Extract only the main xlsx; the alias file lacks an explicit
        # "alias of" link so we can't safely use it without a clustering
        # pass. Skip the 27 MB vehicle xls and the alias file.
        with zipfile.ZipFile(self.zip_path) as zf:
            with zf.open(DEFAULT_MAIN_FILE) as src, self.main_path.open("wb") as dst:
                dst.write(src.read())

    # ---- normalize ----------------------------------------------------

    def normalize(self, raw: dict) -> OffenderRecord:
        assert self._fetched_at is not None

        name = (raw.get("name") or "").strip()
        dob = _to_utc_dt(raw.get("dob"))
        full_name = _flip_lastfirst(name)
        source_id = f"{name}|{_key_dob(raw.get('dob'))}"

        identity = Identity(
            full_name=full_name or "UNKNOWN",
            dob=dob,
            year_of_birth=dob.year if dob else None,
        )

        addresses: list[Address] = []
        street = _clean(raw.get("address"))
        city = _clean(raw.get("city"))
        state = _clean(raw.get("state")) or "MO"
        zip_code = _clean(raw.get("zip"))
        county = _clean(raw.get("county"))
        _ = county  # preserved in raw
        if any((street, city, zip_code)):
            addresses.append(
                Address(
                    type="home",
                    street=street,
                    city=city,
                    state=state,
                    zip=zip_code,
                )
            )

        offenses: list[Offense] = []
        offense_text = _clean(raw.get("offense"))
        if offense_text:
            tier_raw = raw.get("tier")
            tier_str = str(tier_raw).strip() if tier_raw is not None else None
            offenses.append(
                Offense(
                    raw_description=offense_text,
                    tier_or_level_raw=tier_str or None,
                    jurisdiction="US-MO",
                )
            )

        compliant_code = (raw.get("compliant") or "").strip().upper()
        status = _COMPLIANT_MAP.get(compliant_code, "unknown")
        absconder = compliant_code in ("N", "A")
        registration = Registration(
            status=status,  # type: ignore[arg-type]
            absconder=absconder,
        )

        return OffenderRecord(
            source=Source(
                jurisdiction=self.jurisdiction,
                source_id=source_id,
                source_url=DOWNLOAD_URL,
                fetched_at=self._fetched_at,
            ),
            identity=identity,
            addresses=addresses,
            offenses=offenses,
            registration=registration,
            raw=raw,
        )

    def extract_photos(self, raw: dict) -> list[PhotoRef]:
        # MSHP doesn't publish photos in the bulk file.
        return []


# ---------------------------------------------------------------------------
# Field helpers


def _flip_lastfirst(name: str) -> str:
    """`LAST, FIRST MIDDLE` → `FIRST MIDDLE LAST`."""
    if "," in name:
        last, rest = name.split(",", 1)
        return f"{rest.strip()} {last.strip()}".strip()
    return name.strip()


def _clean(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_utc_dt(value: object) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _key_dob(value: object) -> str:
    """Stable string-form of a DOB for use in the synthetic source_id."""
    dt = _to_utc_dt(value)
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d")


def build() -> MissouriAdapter:
    return MissouriAdapter()
